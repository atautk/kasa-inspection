import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from PySide6.QtWidgets import (
    QDialog,
    QVBoxLayout,
    QHBoxLayout,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QPushButton,
    QLineEdit,
    QLabel,
    QFileDialog,
    QMessageBox
)

from ..window_utils import restore_or_center, save_geometry

SETTINGS_KEY = "audit_log_dialog"

# logs/app.log satır formatı: "%(asctime)s [%(levelname)s] %(message)s"
# ve operatör aksiyonları hep "[operatör_adı] ..." ile başlar
# (modules/utils/logger.py + app_logger.info("[%s] ...", operator_name, ...)).
LOG_LINE_PATTERN = re.compile(
    r"^(?P<timestamp>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d{3} "
    r"\[(?P<level>\w+)\] "
    r"(?:\[(?P<operator>[^\]]+)\] )?(?P<message>.*)$"
)


class AuditLogDialog(QDialog):
    """
    logs/app.log dosyasındaki giriş/çıkış ve değişiklik (operatör
    aksiyonu) satırlarını okunabilir bir tabloda gösterir. Bu
    pencereyi kimin açabileceği (yönetici kontrolü) çağıran taraftan
    (ConfiguratorController) yapılır.
    """

    def __init__(self, log_path: Path, parent=None):

        super().__init__(parent)

        self.log_path = log_path

        self.setWindowTitle("Giriş/Çıkış ve Değişiklik Logu")
        restore_or_center(self, SETTINGS_KEY, 900, 600)

        layout = QVBoxLayout(self)

        filter_row = QHBoxLayout()

        filter_row.addWidget(QLabel("Operatör filtrele:"))

        self.filter_input = QLineEdit()
        self.filter_input.textChanged.connect(self.reload)
        filter_row.addWidget(self.filter_input)

        self.refresh_button = QPushButton("Yenile")
        self.refresh_button.clicked.connect(self.reload)
        filter_row.addWidget(self.refresh_button)

        self.export_button = QPushButton("Excel'e Aktar")
        self.export_button.clicked.connect(self._on_export_clicked)
        filter_row.addWidget(self.export_button)

        layout.addLayout(filter_row)

        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(
            ["Zaman", "Operatör", "Mesaj"]
        )
        self.table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.Stretch
        )
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.table)

        self.reload()

    # -------------------------------------------------

    def closeEvent(self, event):

        save_geometry(self, SETTINGS_KEY)

        super().closeEvent(event)

    # -------------------------------------------------
    # Yükleme
    # -------------------------------------------------

    def reload(self):

        entries = self._read_entries()

        query = self.filter_input.text().strip().lower()

        if query:

            entries = [
                entry for entry in entries
                if query in entry[1].lower()
            ]

        entries.reverse()

        self.table.setRowCount(len(entries))

        for row, (timestamp, operator, message) in enumerate(entries):

            for column, value in enumerate(
                [timestamp, operator, message]
            ):
                self.table.setItem(
                    row, column, QTableWidgetItem(value)
                )

    def _read_entries(self):

        if not self.log_path.exists():
            return []

        entries = []

        with open(
            self.log_path, "r", encoding="utf-8", errors="replace"
        ) as f:

            for line in f:

                match = LOG_LINE_PATTERN.match(line.rstrip("\n"))

                if match is None:
                    continue

                operator = match.group("operator") or "-"

                entries.append((
                    match.group("timestamp"),
                    operator,
                    match.group("message")
                ))

        return entries

    # -------------------------------------------------
    # Excel'e Aktar
    # -------------------------------------------------

    def _on_export_clicked(self):

        if self.table.rowCount() == 0:

            QMessageBox.information(
                self, "Bilgi", "Aktarılacak kayıt yok."
            )

            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Excel'e Aktar",
            "giris_cikis_degisiklik_logu.xlsx",
            "Excel Dosyası (*.xlsx)"
        )

        if not path:
            return

        try:

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Log"

            headers = ["Zaman", "Operatör", "Mesaj"]

            for column, header in enumerate(headers, start=1):

                cell = sheet.cell(row=1, column=column, value=header)
                cell.font = Font(bold=True)

            for row in range(self.table.rowCount()):

                for column in range(3):

                    item = self.table.item(row, column)

                    sheet.cell(
                        row=row + 2,
                        column=column + 1,
                        value=item.text() if item else ""
                    )

            sheet.column_dimensions["A"].width = 22
            sheet.column_dimensions["B"].width = 18
            sheet.column_dimensions["C"].width = 70

            workbook.save(path)

        except Exception as e:

            QMessageBox.critical(
                self, "Hata", f"Excel'e aktarma başarısız: {e}"
            )

            return

        QMessageBox.information(
            self,
            "Başarılı",
            f"Log şu dosyaya aktarıldı:\n{path}"
        )
