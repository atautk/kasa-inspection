from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font


class PeriodicReportExporter:
    """
    InspectionLogger.compute_period_stats()'in sonucunu tek sayfalık
    bir özet Excel dosyasına dönüştürür (günlük Telegram raporu için).
    Detaylı kayıt listesi için InspectionLogExporter kullanılır - bu
    sınıf sadece toplam/model/ROI özetini üretir.
    """

    def export(self, stats: dict, destination_path, band_name: str, period_label: str):

        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Özet"

        sheet.append([f"{band_name} - {period_label}"])
        sheet["A1"].font = Font(bold=True, size=14)

        sheet.append([])

        ng_ratio = (
            stats["ng_count"] / stats["total"] * 100
            if stats["total"] else 0.0
        )

        sheet.append(["Toplam Kontrol", stats["total"]])
        sheet.append(["UYGUN", stats["ok_count"]])
        sheet.append(["HATA", stats["ng_count"]])
        sheet.append(["HATA Oranı (%)", round(ng_ratio, 2)])

        sheet.append([])
        sheet.append(["Model", "UYGUN", "HATA"])

        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True)

        for model_name, model_stats in sorted(stats["by_model"].items()):

            sheet.append([
                model_name, model_stats["ok"], model_stats["ng"]
            ])

        sheet.append([])
        sheet.append(["Göz", "UYGUN", "HATA"])

        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True)

        for roi_name, roi_stats in sorted(stats["by_roi"].items()):

            sheet.append([
                roi_name, roi_stats["ok"], roi_stats["ng"]
            ])

        for column in "ABC":
            sheet.column_dimensions[column].width = 24

        workbook.save(destination_path)
