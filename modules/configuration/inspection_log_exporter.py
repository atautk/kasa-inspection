from __future__ import annotations

import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class InspectionLogExporter:

    HEADERS = [
        "ID",
        "Zaman",
        "Model",
        "Sonuç",
        "İncelendi",
        "Orijinal Sonuç",
        "ROI Detayları",
        "Fotoğraf Yolu"
    ]

    OK_FILL = PatternFill(
        start_color="C8FFC8", end_color="C8FFC8", fill_type="solid"
    )
    NG_FILL = PatternFill(
        start_color="FFC8C8", end_color="FFC8C8", fill_type="solid"
    )
    REVIEWED_FILL = PatternFill(
        start_color="FFEBB4", end_color="FFEBB4", fill_type="solid"
    )

    ROI_COLUMN_WIDTH = 60
    IMAGE_COLUMN_WIDTH = 50

    # -------------------------------------------------
    # Excel'e Aktar
    # -------------------------------------------------

    def export(self, inspection_logger, destination_path, band_name=""):

        rows = inspection_logger.fetch_recent(1_000_000)

        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Inspection Log"

        title = f"Inspection Geçmişi - {band_name}" if band_name else "Inspection Geçmişi"

        sheet.append([title])
        sheet.append(self.HEADERS)

        for cell in sheet[2]:
            cell.font = Font(bold=True)

        for row in reversed(rows):

            roi_text = self._format_roi_results(row["roi_results"])

            sheet.append([
                row["id"],
                self._format_timestamp(row["timestamp"]),
                row["model_name"] or "-",
                row["overall_result"],
                "Evet" if row.get("reviewed") else "-",
                row.get("original_result") or "-",
                roi_text,
                row.get("image_path") or "-"
            ])

            excel_row = sheet.max_row

            if row.get("reviewed"):
                fill = self.REVIEWED_FILL
            elif row["overall_result"] == "OK":
                fill = self.OK_FILL
            else:
                fill = self.NG_FILL

            for column in range(1, len(self.HEADERS) + 1):
                sheet.cell(row=excel_row, column=column).fill = fill

        for column_index, header in enumerate(self.HEADERS, start=1):

            sheet.column_dimensions[
                get_column_letter(column_index)
            ].width = max(12, len(header) + 2)

        roi_column = self.HEADERS.index("ROI Detayları") + 1
        image_column = self.HEADERS.index("Fotoğraf Yolu") + 1

        sheet.column_dimensions[
            get_column_letter(roi_column)
        ].width = self.ROI_COLUMN_WIDTH

        sheet.column_dimensions[
            get_column_letter(image_column)
        ].width = self.IMAGE_COLUMN_WIDTH

        workbook.save(destination_path)

    # -------------------------------------------------

    def _format_roi_results(self, roi_results_json: str) -> str:

        try:
            roi_results = json.loads(roi_results_json)
        except Exception:
            return ""

        parts = []

        for name, data in sorted(roi_results.items()):

            status = "OK" if data.get("ok") else "NG"

            parts.append(
                f"{name}: {status} "
                f"({data.get('state')}, beklenen {data.get('expected')})"
            )

        return "; ".join(parts)

    # -------------------------------------------------

    def _format_timestamp(self, iso_text: str) -> str:

        try:

            dt = datetime.fromisoformat(iso_text)

            return dt.astimezone().strftime("%Y-%m-%d %H:%M:%S")

        except Exception:

            return iso_text
