from openpyxl import load_workbook

from modules.configuration.periodic_report_exporter import (
    PeriodicReportExporter
)


def _stats():

    return {
        "total": 10,
        "ok_count": 7,
        "ng_count": 3,
        "by_model": {
            "Clio": {"ok": 5, "ng": 1},
            "Duster": {"ok": 2, "ng": 2}
        },
        "by_roi": {
            "G01": {"ok": 8, "ng": 2},
            "G02": {"ok": 9, "ng": 1}
        }
    }


def test_export_creates_file(tmp_path):

    destination = tmp_path / "rapor.xlsx"

    PeriodicReportExporter().export(
        _stats(), destination, "Test Bandı", "Son 24 Saat Özeti"
    )

    assert destination.exists()


def test_export_contains_totals_and_ratio(tmp_path):

    destination = tmp_path / "rapor.xlsx"

    PeriodicReportExporter().export(
        _stats(), destination, "Test Bandı", "Son 24 Saat Özeti"
    )

    workbook = load_workbook(destination)
    sheet = workbook.active

    values = [
        [cell.value for cell in row]
        for row in sheet.iter_rows()
    ]

    flat_labels = [row[0] for row in values if row and row[0]]

    assert "Toplam Kontrol" in flat_labels
    assert "UYGUN" in flat_labels
    assert "HATA" in flat_labels
    assert "HATA Oranı (%)" in flat_labels

    totals_row = next(
        row for row in values if row and row[0] == "Toplam Kontrol"
    )
    assert totals_row[1] == 10

    ratio_row = next(
        row for row in values if row and row[0] == "HATA Oranı (%)"
    )
    assert ratio_row[1] == 30.0


def test_export_contains_model_and_roi_breakdown(tmp_path):

    destination = tmp_path / "rapor.xlsx"

    PeriodicReportExporter().export(
        _stats(), destination, "Test Bandı", "Son 24 Saat Özeti"
    )

    workbook = load_workbook(destination)
    sheet = workbook.active

    rows = [
        [cell.value for cell in row]
        for row in sheet.iter_rows()
    ]

    assert ["Clio", 5, 1] in rows
    assert ["Duster", 2, 2] in rows
    assert ["G01", 8, 2] in rows
    assert ["G02", 9, 1] in rows


def test_export_handles_zero_total_without_division_error(tmp_path):

    destination = tmp_path / "rapor.xlsx"

    empty_stats = {
        "total": 0, "ok_count": 0, "ng_count": 0,
        "by_model": {}, "by_roi": {}
    }

    PeriodicReportExporter().export(
        empty_stats, destination, "Test Bandı", "Son 24 Saat Özeti"
    )

    workbook = load_workbook(destination)
    sheet = workbook.active

    rows = [[cell.value for cell in row] for row in sheet.iter_rows()]

    ratio_row = next(
        row for row in rows if row and row[0] == "HATA Oranı (%)"
    )
    assert ratio_row[1] == 0.0
